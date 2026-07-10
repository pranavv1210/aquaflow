from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import Client, create_client


DEFAULT_WORKBOOK = Path("water supply details.xlsx")
DEFAULT_REPORT = Path("migration_report.md")
DEFAULT_ORDER_TIME = "09:00:00"
CHUNK_SIZE = 100

EXPENSE_CATEGORY_MAP = {
    "diesel": "Diesel",
    "driver_payment": "Driver Payment",
    "service": "Service",
    "repair": "Repair",
    "police": "Police",
    "tyre": "Tyre",
    "other": "Other",
}


@dataclass
class TableResult:
    table_name: str
    imported: int = 0
    reused: int = 0
    skipped: int = 0
    warnings: list[str] = field(default_factory=list)
    skipped_rows: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


@dataclass
class MigrationReport:
    workbook_path: str
    sheet_names: list[str]
    locations: TableResult = field(default_factory=lambda: TableResult("locations"))
    water_points: TableResult = field(default_factory=lambda: TableResult("water_points"))
    customers: TableResult = field(default_factory=lambda: TableResult("customers"))
    drivers: TableResult = field(default_factory=lambda: TableResult("drivers"))
    vehicles: TableResult = field(default_factory=lambda: TableResult("vehicles"))
    partner_tankers: TableResult = field(default_factory=lambda: TableResult("partner_tankers"))
    orders: TableResult = field(default_factory=lambda: TableResult("orders"))
    expenses: TableResult = field(default_factory=lambda: TableResult("expenses"))
    warnings: list[str] = field(default_factory=list)
    skipped_rows: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    verification_counts: dict[str, int] = field(default_factory=dict)

    def add_warning(self, message: str) -> None:
        self.warnings.append(message)

    def add_error(self, message: str) -> None:
        self.errors.append(message)

    def add_skipped_row(self, message: str) -> None:
        self.skipped_rows.append(message)

    def to_markdown(self) -> str:
        sections = [
            "# AquaFlow Historical Import Report",
            "",
            "## Workbook",
            f"- Source: `{self.workbook_path}`",
            f"- Sheets: {', '.join(f'`{sheet}`' for sheet in self.sheet_names)}",
            "",
            "## Imported Totals",
            f"- Locations Imported: {self.locations.imported}",
            f"- Water Points Imported: {self.water_points.imported}",
            f"- Customers Imported: {self.customers.imported}",
            f"- Drivers Imported: {self.drivers.imported}",
            f"- Vehicles Imported: {self.vehicles.imported}",
            f"- Partner Tankers Imported: {self.partner_tankers.imported}",
            f"- Orders Imported: {self.orders.imported}",
            f"- Expenses Imported: {self.expenses.imported}",
            "",
            "## Warnings",
        ]
        if self.warnings:
            sections.extend(f"- {warning}" for warning in self.warnings)
        else:
            sections.append("- None")

        sections.extend([
            "",
            "## Skipped Rows",
        ])
        if self.skipped_rows:
            sections.extend(f"- {row}" for row in self.skipped_rows)
        else:
            sections.append("- None")

        sections.extend([
            "",
            "## Errors",
        ])
        if self.errors:
            sections.extend(f"- {error}" for error in self.errors)
        else:
            sections.append("- None")

        sections.extend([
            "",
            "## Verification Results",
        ])
        if self.verification_counts:
            for key, value in self.verification_counts.items():
                sections.append(f"- {key}: {value}")
        else:
            sections.append("- Verification not run")

        return "\n".join(sections) + "\n"


def main() -> int:
    args = parse_args()
    load_dotenv(args.env_file)

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    report_path = Path(args.report).expanduser().resolve()
    supabase_url = os.getenv("SUPABASE_URL", "").strip()
    service_role_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "").strip()
    anon_key = os.getenv("SUPABASE_ANON_KEY", "").strip()

    if not supabase_url:
        print("SUPABASE_URL is missing from .env", file=sys.stderr)
        return 1

    if service_role_key:
        supabase_key = service_role_key
        print("Using SUPABASE_SERVICE_ROLE_KEY from .env")
    elif anon_key:
        supabase_key = anon_key
        print("SUPABASE_SERVICE_ROLE_KEY is missing; falling back to SUPABASE_ANON_KEY.")
        print("Writes may fail if the anon key does not have insert permissions.")
    else:
        print("SUPABASE_SERVICE_ROLE_KEY or SUPABASE_ANON_KEY is required.", file=sys.stderr)
        return 1

    print("Reading workbook...")
    workbook = load_workbook(workbook_path, data_only=True)
    sheet_names = workbook.sheetnames

    data_rows = read_data_sheet(workbook_path)
    customer_reference = read_customer_reference(workbook_path)

    print(f"Loaded {len(data_rows)} historical rows from Data sheet")
    print(f"Loaded {len(customer_reference)} customer reference names")

    client = create_client(supabase_url, supabase_key)
    report = MigrationReport(str(workbook_path), sheet_names)

    existing = load_existing_records(client)

    print("Importing Locations...")
    location_map = import_locations(client, data_rows, existing, report)
    print(f"Imported {report.locations.imported} Locations")

    print("Importing Water Points...")
    water_point_map = import_water_points(client, data_rows, existing, report)
    print(f"Imported {report.water_points.imported} Water Points")

    print("Importing Customers...")
    customer_map = import_customers(client, data_rows, customer_reference, existing, location_map, report)
    print(f"Imported {report.customers.imported} Customers")

    print("Importing Drivers...")
    driver_map = import_drivers(client, data_rows, existing, report)
    print(f"Imported {report.drivers.imported} Drivers")

    print("Importing Vehicles...")
    vehicle_map = import_vehicles(client, data_rows, existing, report)
    print(f"Imported {report.vehicles.imported} Vehicles")

    print("Ensuring Expense Categories...")
    expense_category_map = ensure_expense_categories(client)

    print("Importing Orders...")
    import_orders(
        client=client,
        data_rows=data_rows,
        report=report,
        customer_map=customer_map,
        location_map=location_map,
        water_point_map=water_point_map,
        vehicle_map=vehicle_map,
        driver_map=driver_map,
        existing=existing,
    )
    print(f"Imported {report.orders.imported} Orders")

    print("Importing Expenses...")
    import_expenses(
        client=client,
        data_rows=data_rows,
        report=report,
        vehicle_map=vehicle_map,
        driver_map=driver_map,
        expense_category_map=expense_category_map,
        existing=existing,
    )
    print(f"Imported {report.expenses.imported} Expenses")

    print("Verifying counts...")
    report.verification_counts = verify_counts(client)
    for key, value in report.verification_counts.items():
        print(f"{key}: {value}")

    report_text = report.to_markdown()
    report_path.write_text(report_text, encoding="utf-8")
    print(f"Wrote migration report to {report_path}")
    print("Import Complete")

    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import AquaFlow historical Excel data into Supabase.")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Path to the Excel workbook.")
    parser.add_argument("--report", default=str(DEFAULT_REPORT), help="Path to write the migration report.")
    parser.add_argument("--env-file", default=".env", help="Path to the dotenv file.")
    return parser.parse_args()


def read_data_sheet(workbook_path: Path) -> list[dict[str, Any]]:
    dataframe = pd.read_excel(workbook_path, sheet_name="Data", header=1, engine="openpyxl")
    dataframe.columns = [str(column).strip() for column in dataframe.columns]
    rows: list[dict[str, Any]] = []
    for index, record in dataframe.iterrows():
        values = record.to_dict()
        if not any(not is_blank(value) for value in values.values()):
            continue
        row_number = int(index) + 3
        rows.append({"row_number": row_number, **values})
    return rows


def read_customer_reference(workbook_path: Path) -> list[str]:
    dataframe = pd.read_excel(workbook_path, sheet_name="Customer Data", header=None, engine="openpyxl")
    names: list[str] = []
    for _, row in dataframe.iterrows():
        values = [value for value in row.tolist() if not is_blank(value)]
        if not values:
            continue
        name = canonical_display(values[-1])
        if name and name not in names:
            names.append(name)
    return names


def load_existing_records(client: Client) -> dict[str, dict[str, Any]]:
    return {
        "locations": index_by_normalized_name(fetch_all(client, "locations", "id, name"), "name"),
        "water_points": index_by_normalized_name(fetch_all(client, "water_points", "id, name, location_id"), "name"),
        "customers": index_by_normalized_name(fetch_all(client, "customers", "id, display_name, phone, default_location_id, address, notes"), "display_name"),
        "drivers": index_by_normalized_name(fetch_all(client, "drivers", "id, driver_name, phone"), "driver_name"),
        "vehicles": index_by_normalized_name(fetch_all(client, "vehicles", "id, vehicle_name, registration_number, vehicle_type"), "vehicle_name"),
        "partner_tankers": index_by_normalized_name(fetch_all(client, "partner_tankers", "id, owner_name, vehicle_name, registration_number"), "vehicle_name"),
        "expense_categories": index_by_normalized_expense_type(fetch_all(client, "expense_categories", "id, expense_type, category_name")),
        "orders": index_by_signature(fetch_all(client, "orders", "id, order_date, order_time, customer_id, location_id, water_point_id, vehicle_id, driver_id, partner_tanker_id, delivery_type, load_count, amount, paid_amount, payment_status, delivery_status, remarks"), "order"),
        "expenses": index_by_signature(fetch_all(client, "expenses", "id, expense_date, vehicle_id, driver_id, expense_category_id, amount, remarks"), "expense"),
    }


def fetch_all(client: Client, table_name: str, columns: str) -> list[dict[str, Any]]:
    response = client.table(table_name).select(columns).execute()
    return list(response.data or [])


def index_by_normalized_name(rows: Iterable[dict[str, Any]], field_name: str) -> dict[str, dict[str, Any]]:
    indexed: dict[str, dict[str, Any]] = {}
    for row in rows:
        value = row.get(field_name)
        if is_blank(value):
            continue
        key = normalize_name(value)
        indexed.setdefault(key, row)
    return indexed


def index_by_normalized_expense_type(rows: Iterable[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    indexed: dict[str, dict[str, Any]] = {}
    for row in rows:
        value = row.get("expense_type")
        if is_blank(value):
            continue
        indexed.setdefault(normalize_name(value), row)
    return indexed


def index_by_signature(rows: Iterable[dict[str, Any]], table_type: str) -> dict[tuple[Any, ...], dict[str, Any]]:
    indexed: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in rows:
        signature = build_signature_from_db_row(row, table_type)
        indexed.setdefault(signature, row)
    return indexed


def import_locations(client: Client, data_rows: list[dict[str, Any]], existing: dict[str, dict[str, Any]], report: MigrationReport) -> dict[str, str]:
    source_names = collect_unique_display_names(data_rows, "Location")
    new_rows: list[dict[str, Any]] = []
    location_map: dict[str, str] = {}

    for name in source_names:
        key = normalize_name(name)
        existing_row = existing["locations"].get(key)
        if existing_row:
            location_map[key] = existing_row["id"]
            report.locations.reused += 1
            continue
        new_rows.append({"name": name, "notes": None, "is_active": True})

    inserted = insert_rows(client, "locations", new_rows)
    for row in inserted:
        location_map[normalize_name(row["name"])] = row["id"]
    report.locations.imported += len(inserted)
    return location_map


def import_water_points(client: Client, data_rows: list[dict[str, Any]], existing: dict[str, dict[str, Any]], report: MigrationReport) -> dict[str, str]:
    source_names = collect_unique_display_names(data_rows, "Water point")
    new_rows: list[dict[str, Any]] = []
    water_point_map: dict[str, str] = {}

    for name in source_names:
        key = normalize_name(name)
        existing_row = existing["water_points"].get(key)
        if existing_row:
            water_point_map[key] = existing_row["id"]
            report.water_points.reused += 1
            continue
        new_rows.append({"name": name, "location_id": None, "notes": None, "is_active": True})

    inserted = insert_rows(client, "water_points", new_rows)
    for row in inserted:
        water_point_map[normalize_name(row["name"])] = row["id"]
    report.water_points.imported += len(inserted)
    return water_point_map


def import_customers(
    client: Client,
    data_rows: list[dict[str, Any]],
    customer_reference: list[str],
    existing: dict[str, dict[str, Any]],
    location_map: dict[str, str],
    report: MigrationReport,
) -> dict[str, str]:
    customer_locations = build_customer_location_candidates(data_rows)
    source_names = collect_unique_display_names(data_rows, "Customer")
    for name in customer_reference:
        if name not in source_names:
            source_names.append(name)

    new_rows: list[dict[str, Any]] = []
    customer_map: dict[str, str] = {}

    for name in source_names:
        key = normalize_name(name)
        existing_row = existing["customers"].get(key)
        default_location_id = None
        candidate_location = customer_locations.get(key)
        if candidate_location and len(candidate_location) == 1:
            location_name = next(iter(candidate_location))
            default_location_id = location_map.get(location_name)

        if existing_row:
            customer_map[key] = existing_row["id"]
            report.customers.reused += 1
            continue

        new_rows.append(
            {
                "display_name": name,
                "phone": None,
                "default_location_id": default_location_id,
                "address": None,
                "notes": None,
                "is_active": True,
            }
        )

    inserted = insert_rows(client, "customers", new_rows)
    for row in inserted:
        customer_map[normalize_name(row["display_name"])] = row["id"]
    report.customers.imported += len(inserted)
    return customer_map


def import_drivers(client: Client, data_rows: list[dict[str, Any]], existing: dict[str, dict[str, Any]], report: MigrationReport) -> dict[str, str]:
    source_names = collect_unique_display_names(data_rows, "Customer delivery")
    new_rows: list[dict[str, Any]] = []
    driver_map: dict[str, str] = {}

    for name in source_names:
        key = normalize_name(name)
        existing_row = existing["drivers"].get(key)
        if existing_row:
            driver_map[key] = existing_row["id"]
            report.drivers.reused += 1
            continue
        new_rows.append({"driver_name": name, "phone": None, "status": "available", "notes": None, "is_active": True})

    inserted = insert_rows(client, "drivers", new_rows)
    for row in inserted:
        driver_map[normalize_name(row["driver_name"])] = row["id"]
    report.drivers.imported += len(inserted)
    return driver_map


def import_vehicles(client: Client, data_rows: list[dict[str, Any]], existing: dict[str, dict[str, Any]], report: MigrationReport) -> dict[str, str]:
    source_names = collect_unique_display_names(data_rows, "Vehicle")
    type_lookup = {"tractor": "tractor", "canter": "canter"}
    placeholder_reg = {"tractor": "UPDATE-TRACTOR", "canter": "UPDATE-CANTER"}
    new_rows: list[dict[str, Any]] = []
    vehicle_map: dict[str, str] = {}

    for name in source_names:
        key = normalize_name(name)
        existing_row = existing["vehicles"].get(key)
        vehicle_type = type_lookup.get(key, "tractor")
        if existing_row:
            vehicle_map[key] = existing_row["id"]
            report.vehicles.reused += 1
            continue
        new_rows.append(
            {
                "vehicle_name": name,
                "registration_number": placeholder_reg.get(key, f"UPDATE-{name.upper().replace(' ', '-')[:20]}") if key in type_lookup else f"UPDATE-{name.upper().replace(' ', '-')[:20]}",
                "vehicle_type": vehicle_type,
                "status": "available",
                "notes": "Placeholder registration created by historical import tool.",
                "is_active": True,
            }
        )

    inserted = insert_rows(client, "vehicles", new_rows)
    for row in inserted:
        vehicle_map[normalize_name(row["vehicle_name"])] = row["id"]
    report.vehicles.imported += len(inserted)
    return vehicle_map


def ensure_expense_categories(client: Client) -> dict[str, str]:
    payload = [
        {"category_name": "Diesel", "expense_type": "diesel", "description": "Fuel expenses for tankers and business vehicles.", "is_active": True},
        {"category_name": "Driver Payment", "expense_type": "driver_payment", "description": "Payments made to drivers.", "is_active": True},
        {"category_name": "Service", "expense_type": "service", "description": "Scheduled vehicle service and maintenance.", "is_active": True},
        {"category_name": "Repair", "expense_type": "repair", "description": "Vehicle or equipment repair expenses.", "is_active": True},
        {"category_name": "Police", "expense_type": "police", "description": "Police or route-related payments.", "is_active": True},
        {"category_name": "Tyre", "expense_type": "tyre", "description": "Tyre purchase, puncture, and tyre repair expenses.", "is_active": True},
        {"category_name": "Other", "expense_type": "other", "description": "Other business expenses.", "is_active": True},
    ]
    response = client.table("expense_categories").upsert(payload, on_conflict="expense_type").execute()
    rows = list(response.data or [])
    return {normalize_name(row["expense_type"]): row["id"] for row in rows}


def import_orders(
    client: Client,
    data_rows: list[dict[str, Any]],
    report: MigrationReport,
    customer_map: dict[str, str],
    location_map: dict[str, str],
    water_point_map: dict[str, str],
    vehicle_map: dict[str, str],
    driver_map: dict[str, str],
    existing: dict[str, dict[str, Any]],
) -> None:
    customer_defaults = build_customer_defaults(data_rows, "Customer", "Location")
    customer_water_defaults = build_customer_defaults(data_rows, "Customer", "Water point")
    customer_driver_defaults = build_customer_defaults(data_rows, "Customer", "Customer delivery")

    payload: list[dict[str, Any]] = []
    seen_signatures = set(existing["orders"].keys())

    for row in data_rows:
        row_number = row["row_number"]
        customer_name = canonical_display(row.get("Customer"))
        location_name = canonical_display(row.get("Location"))
        water_point_name = canonical_display(row.get("Water point"))
        vehicle_name = canonical_display(row.get("Vehicle"))
        driver_name = canonical_display(row.get("Customer delivery"))
        amount = parse_amount(row.get("Cost"))
        amount_source = "Cost"
        if amount is None:
            amount = parse_amount(row.get("Water Sales"))
            amount_source = "Water Sales"

        if amount is None:
            report.add_skipped_row(f"Row {row_number}: amount missing in both Cost and Water Sales")
            report.orders.skipped += 1
            continue

        if not customer_name or normalize_name(customer_name) not in customer_map:
            report.add_skipped_row(f"Row {row_number}: customer could not be resolved")
            report.orders.skipped += 1
            continue

        customer_key = normalize_name(customer_name)

        if is_blank(location_name):
            location_name = unique_default(customer_defaults.get(customer_key))
            if location_name:
                report.add_warning(f"Row {row_number}: location inferred from customer default")

        if is_blank(water_point_name):
            water_point_name = unique_default(customer_water_defaults.get(customer_key))
            if water_point_name:
                report.add_warning(f"Row {row_number}: water point inferred from customer default")

        if is_blank(driver_name):
            driver_name = unique_default(customer_driver_defaults.get(customer_key))
            if driver_name:
                report.add_warning(f"Row {row_number}: driver inferred from customer default")

        if is_blank(location_name) or is_blank(water_point_name) or is_blank(vehicle_name) or is_blank(driver_name):
            report.add_skipped_row(
                f"Row {row_number}: required reference missing after safe inference"
            )
            report.orders.skipped += 1
            continue

        location_id = location_map.get(normalize_name(location_name))
        water_point_id = water_point_map.get(normalize_name(water_point_name))
        vehicle_id = vehicle_map.get(normalize_name(vehicle_name))
        driver_id = driver_map.get(normalize_name(driver_name))
        customer_id = customer_map.get(customer_key)

        if not all([customer_id, location_id, water_point_id, vehicle_id, driver_id]):
            report.add_skipped_row(f"Row {row_number}: one or more foreign keys could not be resolved")
            report.orders.skipped += 1
            continue

        payment_status = normalize_status(row.get("Payment Status"))
        paid_amount = 0.0
        if payment_status == "paid" or payment_status == "recd":
            payment_status = "paid"
            paid_amount = amount
        elif payment_status == "pending":
            payment_status = "unpaid"
            paid_amount = 0.0
        else:
            report.add_warning(f"Row {row_number}: payment status '{row.get('Payment Status')}' normalized to unpaid")
            payment_status = "unpaid"

        remarks = canonical_display(row.get("Remarks")) or canonical_display(row.get("Remarks 2"))
        order_row = {
            "order_date": to_date(row.get("Date")),
            "order_time": DEFAULT_ORDER_TIME,
            "customer_id": customer_id,
            "location_id": location_id,
            "water_point_id": water_point_id,
            "vehicle_id": vehicle_id,
            "driver_id": driver_id,
            "partner_tanker_id": None,
            "delivery_type": "own_vehicle",
            "load_count": int(parse_load_count(row.get("Load count"))),
            "amount": round(float(amount), 2),
            "paid_amount": round(float(paid_amount), 2),
            "payment_status": payment_status,
            "delivery_status": "delivered",
            "remarks": remarks,
        }
        signature = build_order_signature(order_row)
        if signature in seen_signatures:
            report.orders.reused += 1
            continue
        seen_signatures.add(signature)
        if amount_source == "Water Sales":
            report.add_warning(f"Row {row_number}: amount sourced from Water Sales because Cost was empty")
        payload.append(order_row)

    inserted = insert_rows(client, "orders", payload)
    report.orders.imported += len(inserted)


def import_expenses(
    client: Client,
    data_rows: list[dict[str, Any]],
    report: MigrationReport,
    vehicle_map: dict[str, str],
    driver_map: dict[str, str],
    expense_category_map: dict[str, str],
    existing: dict[str, dict[str, Any]],
) -> None:
    payload: list[dict[str, Any]] = []
    seen_signatures = set(existing["expenses"].keys())

    for row in data_rows:
        row_number = row["row_number"]
        vehicle_name = canonical_display(row.get("Vehicle"))
        driver_name = canonical_display(row.get("Customer delivery"))
        expense_specs = [
            ("Diesel Expenses", "diesel", canonical_display(row.get("Remarks 2"))),
            ("Driver Expenses", "driver_payment", canonical_display(row.get("Remarks 2"))),
            ("Police", "police", canonical_display(row.get("Remarks 2"))),
            ("Vehicle Service and Expenses", infer_service_type(row.get("Remarks 2")), canonical_display(row.get("Remarks 2"))),
        ]

        vehicle_id = vehicle_map.get(normalize_name(vehicle_name)) if vehicle_name else None
        driver_id = driver_map.get(normalize_name(driver_name)) if driver_name else None

        for column_name, expense_type, remarks in expense_specs:
            amount = parse_amount(row.get(column_name))
            if amount is None:
                continue

            category_id = expense_category_map.get(normalize_name(expense_type))
            if not category_id:
                report.add_warning(f"Row {row_number}: expense category '{expense_type}' not found; skipping {column_name}")
                continue

            expense_row = {
                "expense_date": to_date(row.get("Date")),
                "vehicle_id": vehicle_id,
                "driver_id": driver_id,
                "expense_category_id": category_id,
                "amount": round(float(amount), 2),
                "remarks": remarks,
            }
            signature = build_expense_signature(expense_row)
            if signature in seen_signatures:
                report.expenses.reused += 1
                continue
            seen_signatures.add(signature)
            payload.append(expense_row)

    inserted = insert_rows(client, "expenses", payload)
    report.expenses.imported += len(inserted)


def insert_rows(client: Client, table_name: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not rows:
        return []

    inserted_rows: list[dict[str, Any]] = []
    try:
        for start in range(0, len(rows), CHUNK_SIZE):
            batch = rows[start : start + CHUNK_SIZE]
            response = client.table(table_name).insert(batch).execute()
            inserted_rows.extend(list(response.data or []))
        return inserted_rows
    except Exception:
        if inserted_rows:
            rollback_ids = [row["id"] for row in inserted_rows if row.get("id")]
            if rollback_ids:
                client.table(table_name).delete().in_("id", rollback_ids).execute()
        raise


def verify_counts(client: Client) -> dict[str, int]:
    tables = ["locations", "water_points", "customers", "drivers", "vehicles", "orders", "expenses"]
    counts: dict[str, int] = {}
    for table in tables:
        response = client.table(table).select("id").execute()
        counts[table] = len(response.data or [])
    return counts


def collect_unique_display_names(data_rows: list[dict[str, Any]], column_name: str) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for row in data_rows:
        value = canonical_display(row.get(column_name))
        if not value:
            continue
        key = normalize_name(value)
        if key in seen:
            continue
        seen.add(key)
        names.append(value)
    return names


def build_customer_location_candidates(data_rows: list[dict[str, Any]], source_column: str = "Customer", value_column: str = "Location") -> dict[str, set[str]]:
    mapping: dict[str, set[str]] = {}
    for row in data_rows:
        customer = canonical_display(row.get(source_column))
        value = canonical_display(row.get(value_column))
        if not customer or not value:
            continue
        mapping.setdefault(normalize_name(customer), set()).add(value)
    return mapping


def build_customer_defaults(data_rows: list[dict[str, Any]], customer_column: str, value_column: str) -> dict[str, set[str]]:
    return build_customer_location_candidates(data_rows, customer_column, value_column)


def unique_default(values: set[str] | None) -> str | None:
    if values and len(values) == 1:
        return next(iter(values))
    return None


def canonical_display(value: Any) -> str | None:
    if is_blank(value):
        return None
    text = str(value).strip()
    text = collapse_spaces(text)
    return text if text else None


def normalize_name(value: Any) -> str:
    return collapse_spaces(str(value)).casefold()


def collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    if isinstance(value, pd.Timestamp):
        return False
    if isinstance(value, datetime):
        return False
    if isinstance(value, date):
        return False
    return str(value).strip() == "" or str(value).strip().lower() == "nan"


def parse_amount(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def parse_load_count(value: Any) -> int:
    amount = parse_amount(value)
    if amount is None:
        return 0
    return int(amount)


def normalize_status(value: Any) -> str:
    if is_blank(value):
        return "unpaid"
    text = collapse_spaces(str(value)).casefold()
    if text in {"recd", "received", "paid"}:
        return "paid"
    if text in {"pending", "unpaid"}:
        return "pending"
    if text in {"partial", "partially paid"}:
        return "partial"
    return text


def to_date(value: Any) -> str:
    if isinstance(value, pd.Timestamp):
        return value.date().isoformat()
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        parsed = pd.to_datetime(value, errors="coerce")
        if not pd.isna(parsed):
            return parsed.date().isoformat()
    raise ValueError(f"Unable to parse date value: {value!r}")


def infer_service_type(remarks: Any) -> str:
    text = collapse_spaces(str(remarks or "")).casefold()
    if "tyre" in text or "tire" in text:
        return "tyre"
    if "repair" in text or "fix" in text:
        return "repair"
    if text:
        return "service"
    return "service"


def build_order_signature(order_row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        order_row["order_date"],
        order_row["order_time"],
        order_row["customer_id"],
        order_row["location_id"],
        order_row["water_point_id"],
        order_row["vehicle_id"],
        order_row["driver_id"],
        order_row["partner_tanker_id"],
        order_row["delivery_type"],
        order_row["load_count"],
        order_row["amount"],
        order_row["paid_amount"],
        order_row["payment_status"],
        order_row["delivery_status"],
        normalize_name(order_row["remarks"]) if order_row.get("remarks") else "",
    )


def build_expense_signature(expense_row: dict[str, Any]) -> tuple[Any, ...]:
    return (
        expense_row["expense_date"],
        expense_row["vehicle_id"],
        expense_row["driver_id"],
        expense_row["expense_category_id"],
        expense_row["amount"],
        normalize_name(expense_row["remarks"]) if expense_row.get("remarks") else "",
    )


def build_signature_from_db_row(row: dict[str, Any], table_type: str) -> tuple[Any, ...]:
    if table_type == "order":
        return (
            row.get("order_date"),
            row.get("order_time"),
            row.get("customer_id"),
            row.get("location_id"),
            row.get("water_point_id"),
            row.get("vehicle_id"),
            row.get("driver_id"),
            row.get("partner_tanker_id"),
            row.get("delivery_type"),
            int(row.get("load_count") or 0),
            round(float(row.get("amount") or 0), 2),
            round(float(row.get("paid_amount") or 0), 2),
            row.get("payment_status"),
            row.get("delivery_status"),
            normalize_name(row.get("remarks")) if row.get("remarks") else "",
        )
    return (
        row.get("expense_date"),
        row.get("vehicle_id"),
        row.get("driver_id"),
        row.get("expense_category_id"),
        round(float(row.get("amount") or 0), 2),
        normalize_name(row.get("remarks")) if row.get("remarks") else "",
    )


if __name__ == "__main__":
    raise SystemExit(main())